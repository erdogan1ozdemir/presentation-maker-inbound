#!/usr/bin/env python3
"""
veri_tara.py - Veri klasorunu tarar, hangi kaynagin hangi donemi kapsadigini
tespit eder ve secilen moda gore EKSIK listesini uretir.

Neden gerekli: gecmis destelerde en pahali hata sinifi "yanlis dosyayi yanlis
slayta baglamak" oldu. GA4 CSV export'lari basliginda '# Property:' ve
'# Start date:' satirlarini tasir; GSC export'lari sheet adlarindan taninir.
Bu script o basliklari okuyup dosya -> kaynak -> property -> donem eslemesini
kullaniciya gostermeden slayt uretimine gecilmez.

Kullanim:
    python3 veri_tara.py ./veri --mod M1 --donem 2026-06
    python3 veri_tara.py ./veri --mod M2 --donem 2026-Q1 --model ecommerce
    python3 veri_tara.py ./veri --json
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import re
import sys
import warnings
from datetime import date

warnings.filterwarnings("ignore")

# ----------------------------------------------------------------------------
# Kaynak x mod gereklilik matrisi
# ----------------------------------------------------------------------------
# zorunlu: olmadan deste uretilmez.
# opsiyonel: bolum kapatilabilir; kapatilirsa chat'te bildirilir.

REQ = {
    "gsc_seri":        dict(ad="GSC aylik seri (date boyutu, 15-16 ay)",
                            zorunlu=True, kaynak="GSC",
                            not_="Trend grafigi ve T1 tablosu"),
    "gsc_query":       dict(ad="GSC query (donem + karsilastirma donemi)",
                            zorunlu=True, kaynak="GSC",
                            not_="Brand/non-brand kirilimi T2, artis-dusus T3"),
    "gsc_page":        dict(ad="GSC page (donem + karsilastirma donemi)",
                            zorunlu=True, kaynak="GSC",
                            not_="En cok trafik getiren sayfalar T4"),
    "ga4_kanal":       dict(ad="GA4 traffic acquisition (kanal bazli)",
                            zorunlu=True, kaynak="GA4",
                            not_="Kanal YoY/MoM tablosu T6"),
    "ga4_organik":     dict(ad="GA4 organik aylik seri (13-15 ay)",
                            zorunlu=True, kaynak="GA4",
                            not_="Organik session ozeti T7"),
    "ga4_revenue":     dict(ad="GA4 revenue & transaction",
                            zorunlu=False, kaynak="GA4", model="ecommerce",
                            not_="Revenue/transaction tablosu T8"),
    "ga4_urun":        dict(ad="GA4 ecommerce purchases (item bazli)",
                            zorunlu=False, kaynak="GA4", model="ecommerce",
                            not_="Urun funnel T10"),
    "ga4_lead":        dict(ad="GA4 form/lead event",
                            zorunlu=False, kaynak="GA4", model="leadgen",
                            not_="Lead tablosu T11"),
    "ga4_sayfa":       dict(ad="GA4 landing page (sayfa bazli session)",
                            zorunlu=False, kaynak="GA4",
                            not_="Blog/icerik bolumu ve top sayfa slaytlari"),
    "ga4_ai_referral": dict(ad="GA4 AI referral (source/medium filtreli)",
                            zorunlu=False, kaynak="GA4",
                            not_="AI referral trafik T9"),
    "kwp":             dict(ad="Keyword Planner arama hacmi (brand / brand+kategori / non-brand / rakip)",
                            zorunlu=False, kaynak="Keyword Planner",
                            not_="Arama hacmi bolumu T12. Talep-performans "
                                 "ayristirmasi bu veri olmadan yapilamaz"),
    "seomonitor":      dict(ad="SEOmonitor visibility / SoC / AI SoV",
                            zorunlu=False, kaynak="SEOmonitor",
                            not_="Gorunurluk ve rakip bolumu T13/T14"),
    "ahrefs":          dict(ad="Ahrefs siralama dagilimi (ilk 3/10/100)",
                            zorunlu=False, kaynak="Ahrefs",
                            not_="Siralama alinan kelime trendi"),
    "cwv":             dict(ad="CrUX / Core Web Vitals",
                            zorunlu=False, kaynak="CrUX",
                            not_="Core Web Vitals slayti"),
}

# Mod bazli ek gereklilik: karsilastirma donemi sayisi
MOD_KARSILASTIRMA = {
    "M1": ["onceki ay (MoM)", "gecen yil ayni ay (YoY)"],
    "M2": ["onceki ceyrek (QoQ)", "gecen yil ayni ceyrek (Q-YoY)"],
    "M3": ["gecen yil ayni yariyil (H-YoY)"],
    "M4": ["simetrik onceki pencere (before)"],
}

# ----------------------------------------------------------------------------
# Dosya tanima
# ----------------------------------------------------------------------------

PATTERNS = [
    # (anahtar, dosya adi regex, ipucu)
    ("ga4_kanal",       r"traffic[_\- ]acquisition|session.*channel.*group|"
                        r"channel[_\- ]group", "GA4 kanal export'u"),
    ("ga4_organik",     r"organic.*(month|aylik|monthly|seri)|"
                        r"(month|monthly).*organic", "GA4 organik aylik seri"),
    ("ga4_revenue",     r"revenue|monetization|transaction|purchase(?!s?_item)",
                        "GA4 revenue/transaction"),
    ("ga4_urun",        r"item[_\- ]?name|ecommerce[_\- ]purchase|urun|product",
                        "GA4 urun funnel"),
    ("ga4_lead",        r"form[_\- ]?success|lead|generate[_\- ]?lead",
                        "GA4 lead event"),
    ("ga4_ai_referral", r"ai[_\- ]?referral|chatgpt|perplexity|llm[_\- ]?traffic",
                        "GA4 AI referral"),
    ("gsc_seri",        r"performance[_\- ]on[_\- ]search|search[_\- ]?console|"
                        r"gsc", "GSC performans export'u"),
    ("gsc_query",       r"quer(y|ies)|kelime|keyword(?!.*planner)",
                        "GSC query kirilimi"),
    ("gsc_page",        r"page|sayfa|landing", "GSC/GA4 sayfa kirilimi"),
    ("kwp",             r"keyword[_\- ]?planner|arama[_\- ]?hacmi|"
                        r"search[_\- ]?volume|hacim", "Keyword Planner"),
    ("seomonitor",      r"seomonitor|visibility|gorunurluk|share[_\- ]of|sov|soc",
                        "SEOmonitor"),
    ("ahrefs",          r"ahrefs|rank[_\- ]?tracker|organic[_\- ]?keywords",
                        "Ahrefs"),
    ("cwv",             r"crux|core[_\- ]?web|lcp|cls|inp", "CrUX / CWV"),
]

DATE_RX = [
    re.compile(r"(20\d{2})[-_/](\d{1,2})[-_/](\d{1,2})"),
    re.compile(r"(\d{1,2})[.](\d{1,2})[.](20\d{2})"),
]
MONTHS_TR = {"oca": 1, "şub": 2, "sub": 2, "mar": 3, "nis": 4, "may": 5,
             "haz": 6, "tem": 7, "ağu": 8, "agu": 8, "eyl": 9, "eki": 10,
             "kas": 11, "ara": 12}


def _parse_date(s):
    for rx in DATE_RX:
        m = rx.search(s)
        if m:
            g = [int(x) for x in m.groups()]
            try:
                return date(g[0], g[1], g[2]) if g[0] > 1000 \
                    else date(g[2], g[1], g[0])
            except ValueError:
                pass
    return None


def read_csv_header(path, limit=40):
    """GA4 export'larinin '# Property / # Start date' basliklarini okur."""
    meta, hdr = {}, None
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            for i, line in enumerate(f):
                if i > limit:
                    break
                s = line.strip()
                if s.startswith("#"):
                    body = s.lstrip("# ").strip()
                    if ":" in body:
                        k, v = body.split(":", 1)
                        meta[k.strip().lower()] = v.strip()
                elif s and hdr is None:
                    hdr = next(csv.reader([s]), None)
    except Exception:
        pass
    return meta, hdr


def read_xlsx_sheets(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        out = {}
        for name in wb.sheetnames:
            ws = wb[name]
            out[name] = ws.max_row
        wb.close()
        return out
    except Exception:
        return {}


def identify(path):
    """Dosyayi kaynaga esler; property ve donem bilgisini cikarir."""
    base = os.path.basename(path)
    low = base.lower()
    ext = os.path.splitext(low)[1]
    hit = []
    for key, rx, hint in PATTERNS:
        if re.search(rx, low):
            hit.append((key, hint))

    info = dict(dosya=base, yol=path, ext=ext, keys=[k for k, _ in hit],
                ipucu=[h for _, h in hit], property=None, baslangic=None,
                bitis=None, satir=None, sheets=None, uyari=[])

    if ext in (".csv", ".tsv", ".txt"):
        meta, hdr = read_csv_header(path)
        for k in ("property", "property name", "mülk", "site"):
            if k in meta:
                info["property"] = meta[k]
                break
        for k in ("start date", "başlangıç", "baslangic"):
            if k in meta:
                info["baslangic"] = meta[k]
                break
        for k in ("end date", "bitiş", "bitis"):
            if k in meta:
                info["bitis"] = meta[k]
                break
        if hdr:
            info["kolonlar"] = hdr[:12]
            hl = " ".join(c.lower() for c in hdr)
            if ("channel" in hl or "kanal" in hl) \
                    and "ga4_kanal" not in info["keys"]:
                info["keys"].append("ga4_kanal")
            if "landing" in hl and "ga4_sayfa" not in info["keys"]:
                info["keys"].append("ga4_sayfa")
            if ("clicks" in hl or "tıklama" in hl) and "impressions" in hl \
                    and not any(k.startswith("gsc") for k in info["keys"]):
                info["keys"].append("gsc_seri")
        try:
            with open(path, encoding="utf-8-sig", errors="replace") as f:
                info["satir"] = sum(1 for ln in f if ln.strip()
                                    and not ln.startswith("#")) - 1
        except Exception:
            pass
    elif ext in (".xlsx", ".xlsm"):
        sheets = read_xlsx_sheets(path)
        info["sheets"] = sheets
        sl = " ".join(s.lower() for s in sheets)
        if any(w in sl for w in ("quer", "kelime")):
            info["keys"].append("gsc_query")
        if any(w in sl for w in ("page", "sayfa")):
            info["keys"].append("gsc_page")
        if any(w in sl for w in ("date", "tarih")):
            info["keys"].append("gsc_seri")

    if not info["baslangic"]:
        d = _parse_date(base)
        if d:
            # Dosya adindaki tarih genelde EXPORT tarihidir, veri donemi degil.
            # Ikisini karistirmak yanlis donem etiketine yol acar.
            info["export_tarihi"] = d.isoformat()
            info["uyari"].append(
                "veri donemi dosyadan okunamadi; dosya adindaki tarih export "
                "tarihi. Hangi donemi kapsadigi kullaniciya sorulmali")
    for tr, mn in MONTHS_TR.items():
        if re.search(r"\b" + tr + r"[a-zçğıöşü]*[\s_\-']?(2\d)?", low):
            info.setdefault("ay_ipucu", []).append(mn)
            break

    # Property alani GA4 export'unu ele verir: bu dosyaya gsc_* etiketi
    # yapistirmak "yanlis dosyayi yanlis slayta baglama" hatasinin kaynagi.
    if (info.get("property") or "").upper().find("GA4") >= 0:
        info["keys"] = [k for k in info["keys"] if not k.startswith("gsc_")]
    info["keys"] = list(dict.fromkeys(info["keys"]))
    if not info["keys"]:
        info["uyari"].append("kaynak tanimlanamadi - hangi veriye ait oldugu "
                             "kullaniciya sorulmali")
    return info


# ----------------------------------------------------------------------------

def scan(folder):
    files = []
    for p in sorted(glob.glob(os.path.join(folder, "**", "*"), recursive=True)):
        if os.path.isdir(p):
            continue
        if os.path.splitext(p)[1].lower() not in (
                ".csv", ".tsv", ".txt", ".xlsx", ".xlsm", ".json", ".md",
                ".png", ".jpg", ".jpeg"):
            continue
        if os.path.basename(p).startswith("~$") or \
                os.path.basename(p).startswith("."):
            continue
        files.append(identify(p))
    return files


def coverage(files, mod, model):
    have = {}
    for f in files:
        for k in f["keys"]:
            have.setdefault(k, []).append(f["dosya"])

    var, eksik_zorunlu, eksik_ops = [], [], []
    for key, spec in REQ.items():
        if spec.get("model") and spec["model"] != model:
            continue
        if key in have:
            var.append((key, spec, have[key]))
        elif spec["zorunlu"]:
            eksik_zorunlu.append((key, spec))
        else:
            eksik_ops.append((key, spec))
    bilinmeyen = [f for f in files if not f["keys"]]
    return var, eksik_zorunlu, eksik_ops, bilinmeyen


def main():
    ap = argparse.ArgumentParser(description="Veri klasoru tarama ve eksik listesi")
    ap.add_argument("folder")
    ap.add_argument("--mod", default="M1", choices=["M1", "M2", "M3", "M4"])
    ap.add_argument("--donem", default=None, help="orn. 2026-06 / 2026-Q1")
    ap.add_argument("--model", default="ecommerce",
                    choices=["ecommerce", "leadgen", "subscription"])
    ap.add_argument("--json", action="store_true")
    a = ap.parse_args()

    if not os.path.isdir(a.folder):
        print(f"Klasor bulunamadi: {a.folder}")
        return 1

    files = scan(a.folder)
    var, ez, eo, bilinmeyen = coverage(files, a.mod, a.model)

    if a.json:
        print(json.dumps(dict(dosyalar=files,
                              var=[k for k, _s, _f in var],
                              eksik_zorunlu=[k for k, _ in ez],
                              eksik_opsiyonel=[k for k, _ in eo],
                              tanimlanamayan=[f["dosya"] for f in bilinmeyen]),
                         ensure_ascii=False, indent=1))
        return 0

    print(f"VERI TARAMASI  |  klasor: {a.folder}  |  mod: {a.mod}"
          f"{'  |  donem: ' + a.donem if a.donem else ''}  |  is modeli: {a.model}")
    print("=" * 78)

    print(f"\nTANINAN DOSYALAR ({len(files) - len(bilinmeyen)}/{len(files)})")
    for f in files:
        if not f["keys"]:
            continue
        tag = ", ".join(f["keys"])
        print(f"  · {f['dosya']}")
        print(f"      kaynak : {tag}")
        if f.get("property"):
            print(f"      property: {f['property']}")
        if f.get("baslangic") or f.get("bitis"):
            print(f"      donem  : {f.get('baslangic','?')} - {f.get('bitis','?')}")
        elif f.get("export_tarihi"):
            print(f"      export : {f['export_tarihi']} (veri donemi BILINMIYOR)")
        for u in f.get("uyari", []):
            print(f"      ! {u}")
        if f.get("satir") is not None:
            print(f"      satir  : {f['satir']}")
        if f.get("sheets"):
            print(f"      sheet  : {', '.join(list(f['sheets'])[:8])}")

    # Ayni kaynaga esleсen, donemi belirsiz mukerrer dosyalar: gercek destelerde
    # 6 adet ayni adli GSC export'u farkli donemleri tasiyordu ve hangisinin
    # hangi doneme ait oldugu dosyadan anlasilmiyordu.
    belirsiz = [f for f in files
                if f["keys"] and not (f.get("baslangic") or f.get("bitis"))]
    if len(belirsiz) > 1:
        print(f"\n{'!'*78}")
        print(f"DONEMI BELIRSIZ MUKERRER DOSYA ({len(belirsiz)})")
        print("Ayni kaynaga eslesen ve icinden donem okunamayan dosyalar. Hangi")
        print("dosyanin hangi donemi kapsadigi TEYIT EDILMEDEN slayta baglanmaz:")
        for f in belirsiz:
            print(f"  · {f['dosya']}  ->  {', '.join(f['keys'])}")

    if bilinmeyen:
        print(f"\nTANIMLANAMAYAN ({len(bilinmeyen)}) - kullaniciya sorulmali")
        for f in bilinmeyen:
            print(f"  · {f['dosya']}")

    if ez:
        print(f"\n{'!'*78}\nEKSIK - ZORUNLU ({len(ez)}): bunlar gelmeden deste uretilmez")
        for k, s in ez:
            print(f"  · {s['ad']}")
            print(f"      kaynak: {s['kaynak']} | kullanim: {s['not_']}")
        for c in MOD_KARSILASTIRMA[a.mod]:
            print(f"  · Karsilastirma donemi: {c}")

    if eo:
        print(f"\nEKSIK - OPSIYONEL ({len(eo)}): gelmezse ilgili bolum destede yer almaz")
        for k, s in eo:
            print(f"  · {s['ad']}")
            print(f"      kaynak: {s['kaynak']} | etkilenen: {s['not_']}")

    print("\n" + "=" * 78)
    if ez:
        print("SONUC: zorunlu veri eksik. Eksik export'lar istenmeden slayt "
              "uretimine gecilmez.")
        print("Opsiyonel eksikler icin: veri saglanamiyorsa ilgili bolum "
              "destedan komple cikarilir ve bu chat'te ONEMLI olarak belirtilir.")
        return 1
    print("SONUC: zorunlu veri tam. Opsiyonel eksikler kullaniciyla teyit "
          "edilip ilgili bolumler kapatilabilir.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
