# -*- coding: utf-8 -*-
"""Game+ GA4 verisi - kanal × ay × yıl × landing page.

Kaynak: "Free-form 1.xlsx" (Cloud Gaming - GA4, 20250701-20260731).
Yıl kolonu eklenmiş sürüm; önceki export'ta Month yıl taşımadığı için
Temmuz 2025 ile Temmuz 2026 birleşiyordu.

Hijyen: landing page listesinde tarayıcı/tarama aracı kaynaklı bozuk yollar
bulunuyor (ör. içinde tırnak, <> ve rastgele token taşıyan varyantlar).
Bunlar ayrı sayfa gibi göründüğü için ayıklanır.
"""
import re
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path("/Users/Erdo/Downloads/Free-form 1.xlsx")

# tarama aracı artığı: tırnak, açılı parantez, & veya çok uzun rastgele token
KIRLI = re.compile(r"""["'<>&]|/[a-z0-9]{10}(?:/|$)""")


def yukle(path=XLSX):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    hdr, rows = None, []
    for r in ws.iter_rows(values_only=True):
        v = list(r)
        if hdr is None:
            if v and v[0] and str(v[0]).lower().startswith("session default"):
                hdr = [str(x) if x is not None else "" for x in v]
            continue
        if not v[0]:
            continue
        rows.append(v)
    ix = {}
    for i, n in enumerate(hdr):
        if n and n not in ix:
            ix[n] = i
    out = []
    for v in rows:
        lp = str(v[ix["Landing page"]] or "")
        out.append(dict(
            kanal=str(v[ix["Session default channel group"]]),
            ay=int(float(v[ix["Month"]])),
            yil=int(float(v[ix["Year"]])),
            lp=lp,
            kirli=bool(KIRLI.search(lp)),
            sure=float(v[ix["Average session duration"]] or 0),
            session=int(float(v[ix["Sessions"]] or 0)),
            engaged=int(float(v[ix["Engaged sessions"]] or 0)),
            bounce=float(v[ix["Bounce rate"]] or 0),
        ))
    return out


TR = {1: "Oca", 2: "Şub", 3: "Mar", 4: "Nis", 5: "May", 6: "Haz",
      7: "Tem", 8: "Ağu", 9: "Eyl", 10: "Eki", 11: "Kas", 12: "Ara"}


def ym(r):
    return r["yil"] * 100 + r["ay"]


def etiket(y):
    return f"{TR[y % 100]}'{str(y // 100)[2:]}"


if __name__ == "__main__":
    d = yukle()
    temiz = [r for r in d if not r["kirli"]]
    print(f"satır: {len(d)}  ·  temiz: {len(temiz)}  ·  ayıklanan: {len(d)-len(temiz)}")
    print(f"ayıklanan session: {sum(r['session'] for r in d if r['kirli']):,}")

    aylar = sorted({ym(r) for r in temiz})
    print(f"\ndönem: {len(aylar)} ay · {etiket(aylar[0])} - {etiket(aylar[-1])}")

    print("\nTemmuz ayrımı kontrolü (Organic Search):")
    for y in (202507, 202607):
        s = sum(r["session"] for r in temiz if ym(r) == y and r["kanal"] == "Organic Search")
        print(f"  {etiket(y)}: {s:,} session")

    print("\nAylık toplam session (tüm kanallar):")
    for y in aylar:
        s = sum(r["session"] for r in temiz if ym(r) == y)
        o = sum(r["session"] for r in temiz if ym(r) == y and r["kanal"] == "Organic Search")
        print(f"  {etiket(y)}  toplam {s:>9,}   organic {o:>8,}")
