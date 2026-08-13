# -*- coding: utf-8 -*-
"""Game+ ek veri katmanı: haftalık SSR serisi, segment pozisyonları,
GA4 organik kırılımı, AI kaynaklı trafik segmenti, oyun kategorisi GSC clickleri.
"""
import collections
import re
from pathlib import Path

from openpyxl import load_workbook

TR = {1: "Oca", 2: "Şub", 3: "Mar", 4: "Nis", 5: "May", 6: "Haz",
      7: "Tem", 8: "Ağu", 9: "Eyl", 10: "Eki", 11: "Kas", 12: "Ara"}

TOOLS = Path("/Users/Erdo/.claude/projects/-Users-Erdo-Desktop-Claude-Projects-sunum--rnekleri"
             "/0207f43a-132e-425b-9b0f-f6ad91b2f665/tool-results")
TOPLAM_GUNLUK = TOOLS / "mcp-gsc-get_advanced_search_analytics-1786636400207.txt"
OYUN_SAYFA = TOOLS / "mcp-gsc-get_advanced_search_analytics-1786642696119.txt"

GUN = re.compile(r"^(\d{4}-\d{2}-\d{2})\s*\|\s*(\w+)\s*\|\s*(\d+)\s*\|\s*(\d+)\s*\|\s*"
                 r"([\d.]+)%\s*\|\s*([\d.]+)\s*$")
SAYFA = re.compile(r"^(\S+)\s*\|\s*(\d{4}-\d{2}-\d{2})\s*\|\s*(\d+)\s*\|\s*(\d+)\s*\|\s*"
                   r"([\d.]+)%\s*\|\s*([\d.]+)\s*$")


# ----------------------------------------------------------------- haftalık
def haftalik():
    """SSR grafiği için ISO hafta bazında click / impression / pozisyon."""
    import datetime as dt
    h = collections.defaultdict(lambda: [0, 0, 0.0])
    for line in TOPLAM_GUNLUK.read_text(encoding="utf-8").splitlines():
        m = GUN.match(line.strip())
        if not m:
            continue
        t = dt.date.fromisoformat(m.group(1))
        # haftanın pazartesisi
        pzt = t - dt.timedelta(days=t.weekday())
        a = h[pzt]
        a[0] += int(m.group(3))
        a[1] += int(m.group(4))
        a[2] += int(m.group(4)) * float(m.group(6))
    out = []
    for pzt in sorted(h):
        c, i, wp = h[pzt]
        out.append(dict(hafta=pzt, click=c, impr=i, poz=wp / i if i else 0))
    return out


# --------------------------------------------------------- oyun kategorileri
def oyun_gsc():
    """/gfn/oyunlar/ alt kategorileri: ay -> (click, impression)"""
    agg = collections.defaultdict(lambda: collections.defaultdict(lambda: [0, 0]))
    for line in OYUN_SAYFA.read_text(encoding="utf-8").splitlines():
        m = SAYFA.match(line.strip())
        if not m:
            continue
        lp = m.group(1).split("gameplus.com.tr")[-1].split("?")[0].rstrip("/")
        ay = m.group(2)[:7]
        agg[lp][ay][0] += int(m.group(3))
        agg[lp][ay][1] += int(m.group(4))
    return agg


# --------------------------------------------------------------- AI segmenti
def ai_segment(path="/Users/Erdo/Downloads/Free-form 1 (1).xlsx"):
    """GA4 AI kaynaklı trafik segmenti: yil*100+ay -> session (13 ay)."""
    ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    hdr, out = None, {}
    for r in ws.iter_rows(values_only=True):
        v = list(r)
        if hdr is None:
            if v and v[0] and str(v[0]).strip() == "Month":
                hdr = v
            continue
        if v[0] is None or v[1] is None:
            continue
        out[int(v[1]) * 100 + int(v[0])] = int(v[3])
    return out


def etiket(ym):
    return f"{TR[ym % 100]}'{str(ym // 100)[2:]}"


if __name__ == "__main__":
    hs = haftalik()
    print(f"haftalık seri: {len(hs)} hafta · {hs[0]['hafta']} - {hs[-1]['hafta']}")
    import datetime as dt
    gecis = dt.date(2026, 2, 9)          # 8-14 Şubat haftasının pazartesisi
    idx = next(i for i, h in enumerate(hs) if h["hafta"] == gecis)
    print(f"  geçiş haftası index {idx} ({gecis})")
    print(f"  öncesi 8 hafta ort. click {sum(h['click'] for h in hs[idx-8:idx])/8:,.0f}"
          f" · sonrası 8 hafta ort. {sum(h['click'] for h in hs[idx+1:idx+9])/8:,.0f}")

    a = ai_segment()
    print(f"\nAI segmenti: {len(a)} ay · "
          + "  ".join(f"{etiket(y)} {a[y]}" for y in sorted(a)))

    o = oyun_gsc()
    print(f"\noyun kategorisi (GSC): {len(o)} sayfa")
